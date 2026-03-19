"""
Kopiert die befüllte Excel-Quelle ins Repo und exportiert ``lp_import_template`` nach CSV.

Standardquelle: %USERPROFILE%\\Documents\\dlg_2025_lp_import_filled_sourced.xlsx
Ziel:
  - data/reference/dlg_2025_lp_import_filled_sourced.xlsx
  - data/reference/dlg_2025_lp_import_filled_sourced.csv

Aufruf (aus Ordner rationsoptimierung):
  python scripts/import_dlg_lp_filled_from_xlsx.py
  python scripts/import_dlg_lp_filled_from_xlsx.py "C:\\pfad\\datei.xlsx" [blattname]
"""
from __future__ import annotations

import csv
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parents[1]
_DEFAULT_SRC = Path.home() / "Documents" / "dlg_2025_lp_import_filled_sourced.xlsx"
_OUT_XLSX = _ROOT / "data" / "reference" / "dlg_2025_lp_import_filled_sourced.xlsx"
_OUT_CSV = _ROOT / "data" / "reference" / "dlg_2025_lp_import_filled_sourced.csv"
_DEFAULT_SHEETS = ("lp_import_filled", "lp_import_template")


def main() -> None:
    src = Path(sys.argv[1]) if len(sys.argv) > 1 else _DEFAULT_SRC
    sheet_arg = sys.argv[2] if len(sys.argv) > 2 else None

    if not src.is_file():
        raise SystemExit(f"Quelldatei nicht gefunden: {src}")

    _OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, _OUT_XLSX)
    print(f"Kopiert: {src} -> {_OUT_XLSX}")

    wb = load_workbook(src, read_only=True, data_only=True)
    if sheet_arg:
        sheet = sheet_arg
    else:
        sheet = next((s for s in _DEFAULT_SHEETS if s in wb.sheetnames), None)
        if sheet is None:
            sheet = wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Blatt '{sheet}' fehlt. Vorhanden: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise SystemExit("Leeres Blatt")

    with _OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for row in rows:
            if row is None:
                continue
            line = ["" if c is None else c for c in row]
            if not any(str(x).strip() for x in line if x is not None):
                continue
            w.writerow(line)

    print(f"CSV exportiert: {_OUT_CSV} (Blatt '{sheet}', {len(rows)} Zeilen roh)")


if __name__ == "__main__":
    main()
