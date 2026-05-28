"""
Central Export Mikroservice
POST /api/v1/export/list – stream CSV or XLSX for entity lists (debtors, creditors,
chart_of_accounts, dunning_runs, audit_log, users). Tenant-scoped.
"""

from __future__ import annotations

import csv
import io
import json
from datetime import date
from typing import Any, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.core.database import get_db
from app.core.tenant import get_tenant_id


from app.api.v1.schemas.base import BaseSchema
from pydantic import ConfigDict as _ConfigDict


class CompatFlexOut(BaseSchema):
    model_config = _ConfigDict(extra="allow")


router = APIRouter()

# Supported entity identifiers for export
SUPPORTED_ENTITIES = [
    "debtors",
    "creditors",
    "chart_of_accounts",
    "dunning_runs",
    "audit_log",
    "users",
    "futtermittel_einzel",
    "futtermittel_misch",
    "futtermittel_chargen",
    "superglue_domain_rollouts",
]


class ExportListFilter(BaseModel):
    """Optional filter for export (e.g. date_from, tenant_id)."""
    date_from: Optional[str] = None
    date_to: Optional[str] = None
    tenant_id: Optional[str] = None
    entity_type: Optional[str] = None
    entity_id: Optional[str] = None
    user_id: Optional[str] = None
    action: Optional[str] = None


class ExportListRequest(BaseModel):
    """Request body for POST /export/list."""
    entity: str = Field(..., description="Entity to export: debtors, creditors, chart_of_accounts, dunning_runs, audit_log, users")
    format: str = Field(default="csv", description="Export format: csv | xlsx")
    filter: Optional[ExportListFilter] = Field(default=None, description="Optional filters")


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (date,)):
        return v.isoformat()
    if hasattr(v, "isoformat"):
        return v.isoformat()
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


# ---------- Entity export functions (tenant-scoped) ----------

def _export_debtors(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    effective_tenant = (filt.tenant_id if filt else None) or tenant_id
    q = text("""
        SELECT id, tenant_id, debitor_number, name, address, payment_terms, credit_limit,
               is_active, created_at, updated_at
        FROM domain_erp.debitors
        WHERE tenant_id = :tenant_id
        ORDER BY debitor_number
    """)
    rows = db.execute(q, {"tenant_id": effective_tenant}).fetchall()
    _cols = ["id", "tenant_id", "debtor_number", "company_name", "address", "payment_terms", "credit_limit", "is_active", "created_at", "updated_at"]  # noqa: F841
    out = []
    for row in rows:
        addr = row[4]
        if isinstance(addr, str):
            try:
                addr = json.loads(addr) if addr else {}
            except Exception:
                addr = {}
        out.append({
            "id": _safe_str(row[0]),
            "tenant_id": _safe_str(row[1]),
            "debtor_number": _safe_str(row[2]),
            "company_name": _safe_str(row[3]),
            "address": json.dumps(addr, ensure_ascii=False) if isinstance(addr, dict) else _safe_str(addr),
            "payment_terms": _safe_str(row[5]),
            "credit_limit": _safe_str(row[6]),
            "is_active": _safe_str(row[7]),
            "created_at": _safe_str(row[8]),
            "updated_at": _safe_str(row[9]),
        })
    return out


def _export_creditors(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    # Export creditor open items from domain_erp.offene_posten (konto_typ = kreditoren)
    effective_tenant = (filt.tenant_id if filt else None) or tenant_id
    q = text("""
        SELECT id, tenant_id, konto_nr, konto_name, konto_typ, op_status, rechnungsnr, rechnungsdatum, faelligkeit,
               op_betrag, saldo, offen, op_text, waehrung, lieferant_id, lieferant_name, skonto_prozent, skonto_bis,
               mahn_stufe, zahlbar, created_at, updated_at
        FROM domain_erp.offene_posten
        WHERE tenant_id = :tenant_id AND konto_typ = 'kreditoren'
        ORDER BY lieferant_name, faelligkeit
    """)
    try:
        rows = db.execute(q, {"tenant_id": effective_tenant}).fetchall()
    except Exception:
        # Fallback if schema/table differs: return empty list
        return []
    cols = ["id", "tenant_id", "konto_nr", "konto_name", "konto_typ", "op_status", "rechnungsnr", "rechnungsdatum",
            "faelligkeit", "op_betrag", "saldo", "offen", "op_text", "waehrung", "lieferant_id", "lieferant_name",
            "skonto_prozent", "skonto_bis", "mahn_stufe", "zahlbar", "created_at", "updated_at"]
    return [dict(zip(cols, [_safe_str(c) for c in row])) for row in rows]


def _export_chart_of_accounts(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    effective_tenant = (filt.tenant_id if filt else None) or tenant_id
    q = text("""
        SELECT id, tenant_id, account_number, account_name, account_type, category, subcategory,
               description, is_summary, balance, last_transaction_date, is_active, created_at, updated_at
        FROM domain_erp.chart_of_accounts
        WHERE tenant_id = :tenant_id AND (is_active IS NULL OR is_active = true)
        ORDER BY account_number
    """)
    rows = db.execute(q, {"tenant_id": effective_tenant}).fetchall()
    cols = ["id", "tenant_id", "account_number", "account_name", "account_type", "category", "subcategory",
            "description", "is_summary", "balance", "last_transaction_date", "is_active", "created_at", "updated_at"]
    return [dict(zip(cols, [_safe_str(c) for c in row])) for row in rows]


def _export_dunning_runs(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    effective_tenant = (filt.tenant_id if filt else None) or tenant_id
    q = text("""
        SELECT id, op_id, debtor_id, dunning_level, dunning_date, due_date, open_amount, dunning_fee,
               interest, total_amount, payment_deadline, status, sent_date, payment_date, notes, created_at, updated_at
        FROM domain_erp.dunning_notices
        WHERE tenant_id = :tenant_id
        ORDER BY dunning_date DESC, dunning_level DESC
    """)
    rows = db.execute(q, {"tenant_id": effective_tenant}).fetchall()
    cols = ["id", "op_id", "debtor_id", "dunning_level", "dunning_date", "due_date", "open_amount", "dunning_fee",
            "interest", "total_amount", "payment_deadline", "status", "sent_date", "payment_date", "notes", "created_at", "updated_at"]
    return [dict(zip(cols, [_safe_str(c) for c in row])) for row in rows]


def _export_audit_log(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    from app.infrastructure.models import AuditLog

    effective_tenant = (filt.tenant_id if filt else None) or tenant_id
    query = db.query(AuditLog).filter(AuditLog.tenant_id == effective_tenant)
    if filt:
        if filt.entity_type:
            query = query.filter(AuditLog.entity_type == filt.entity_type)
        if filt.entity_id:
            query = query.filter(AuditLog.entity_id == filt.entity_id)
        if filt.user_id:
            query = query.filter(AuditLog.user_id == filt.user_id)
        if filt.action:
            query = query.filter(AuditLog.action == filt.action)
    logs = query.order_by(AuditLog.timestamp.desc()).limit(10_000).all()
    out = []
    for log in logs:
        out.append({
            "id": _safe_str(log.id),
            "timestamp": _safe_str(log.timestamp),
            "user_id": _safe_str(log.user_id),
            "user_email": _safe_str(log.user_email),
            "tenant_id": _safe_str(log.tenant_id),
            "action": _safe_str(log.action),
            "entity_type": _safe_str(log.entity_type),
            "entity_id": _safe_str(log.entity_id),
            "changes": json.dumps(log.changes or {}, ensure_ascii=False),
            "ip_address": _safe_str(log.ip_address),
            "correlation_id": _safe_str(log.correlation_id),
        })
    return out


def _export_users(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    effective_tenant = (filt.tenant_id if filt else None) or tenant_id
    q = text("""
        SELECT id, keycloak_id, username, email, first_name, last_name, roles, tenant_id, is_active, last_login, created_at, updated_at
        FROM domain_shared.users
        WHERE tenant_id = :tenant_id AND (deleted_at IS NULL)
        ORDER BY username
    """)
    rows = db.execute(q, {"tenant_id": effective_tenant}).fetchall()
    cols = ["id", "keycloak_id", "username", "email", "first_name", "last_name", "roles", "tenant_id", "is_active", "last_login", "created_at", "updated_at"]
    return [dict(zip(cols, [_safe_str(c) for c in row])) for row in rows]


def _export_futtermittel_einzel(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    """Export Einzelfuttermittel aus domain_inventory.articles."""
    q = text("""
        SELECT id, article_number, name, category, unit,
               sales_price, purchase_price, hersteller,
               analyse_protein, analyse_feuchtigkeit, naehrwertangaben,
               chargenpflicht, gmp_plus_relevanz, created_at
        FROM domain_inventory.articles
        WHERE is_active = true
        ORDER BY name
    """)
    try:
        rows = db.execute(q).fetchall()
    except Exception:
        return []
    out = []
    for row in rows:
        out.append({
            "id": _safe_str(row[0]),
            "artikelnummer": _safe_str(row[1]),
            "name": _safe_str(row[2]),
            "kategorie": _safe_str(row[3]),
            "einheit": _safe_str(row[4]),
            "vk_preis": _safe_str(row[5]),
            "ek_preis": _safe_str(row[6]),
            "hersteller": _safe_str(row[7]),
            "protein_pct": _safe_str(row[8]),
            "feuchtigkeit_pct": _safe_str(row[9]),
            "naehrwertangaben": _safe_str(row[10]),
            "chargenpflicht": _safe_str(row[11]),
            "gmp_plus": _safe_str(row[12]),
            "erstellt_am": _safe_str(row[13]),
        })
    return out


def _export_futtermittel_misch(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    """Export Mischfuttermittel aus domain_inventory.articles (custom_properties)."""
    q = text("""
        SELECT id, article_number, name, category, unit,
               sales_price, purchase_price, hersteller,
               custom_properties, naehrwertangaben,
               chargenpflicht, gmp_plus_relevanz, created_at
        FROM domain_inventory.articles
        WHERE is_active = true
        ORDER BY name
    """)
    try:
        rows = db.execute(q).fetchall()
    except Exception:
        return []
    out = []
    for row in rows:
        props = row[8]
        if isinstance(props, str):
            try:
                props = json.loads(props)
            except Exception:
                props = {}
        props = props or {}
        out.append({
            "id": _safe_str(row[0]),
            "artikelnummer": _safe_str(row[1]),
            "name": _safe_str(row[2]),
            "kategorie": _safe_str(row[3]),
            "einheit": _safe_str(row[4]),
            "vk_preis": _safe_str(row[5]),
            "ek_preis": _safe_str(row[6]),
            "hersteller": _safe_str(row[7]),
            "tierart": _safe_str(props.get("tierart", "")),
            "phase": _safe_str(props.get("phase", "")),
            "komponenten_anzahl": _safe_str(props.get("komponenten", "")),
            "rohfett_pct": _safe_str(props.get("rohfett", "")),
            "rohfaser_pct": _safe_str(props.get("rohfaser", "")),
            "energie_mj_kg": _safe_str(props.get("energie", "")),
            "naehrwertangaben": _safe_str(row[9]),
            "chargenpflicht": _safe_str(row[10]),
            "gmp_plus": _safe_str(row[11]),
            "erstellt_am": _safe_str(row[12]),
        })
    return out


def _export_futtermittel_chargen(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    """Export Futtermittel-Chargen aus domain_ops.ops_chargen."""
    q = text("""
        SELECT id, chargen_id, artikel, menge, lagerort, status,
               qualitaetsstatus, eingang, mhd, herkunft, bemerkungen
        FROM domain_ops.ops_chargen
        ORDER BY eingang DESC
    """)
    try:
        rows = db.execute(q).fetchall()
    except Exception:
        return []
    cols = ["id", "chargen_id", "artikel", "menge_kg", "lagerort", "status",
            "qualitaetsstatus", "eingang", "mhd", "herkunft", "bemerkungen"]
    return [dict(zip(cols, [_safe_str(c) for c in row])) for row in rows]


def _export_superglue_domain_rollouts(db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    from app.integrations.services.superglue_domain_rollouts import build_superglue_domain_rollout_summary

    summary = build_superglue_domain_rollout_summary((filt.tenant_id if filt else None) or tenant_id)
    rows: list[dict] = []
    for domain in summary["domains"]:
        for connector in domain["connectors"]:
            rows.append(
                {
                    "tenant_id": summary["tenant_id"],
                    "domain_key": domain["domain_key"],
                    "connector_key": connector["connector_key"],
                    "tool_id": connector["tool_id"],
                    "valeo_contract_id": connector["valeo_contract_id"],
                    "execution_modes": ",".join(connector["execution_modes"]),
                    "system_url": connector["system_url"],
                }
            )
    return rows


def _get_entity_data(entity: str, db: Session, tenant_id: str, filt: Optional[ExportListFilter]) -> list[dict]:
    handlers = {
        "debtors": _export_debtors,
        "creditors": _export_creditors,
        "chart_of_accounts": _export_chart_of_accounts,
        "dunning_runs": _export_dunning_runs,
        "audit_log": _export_audit_log,
        "users": _export_users,
        "futtermittel_einzel": _export_futtermittel_einzel,
        "futtermittel_misch": _export_futtermittel_misch,
        "futtermittel_chargen": _export_futtermittel_chargen,
        "superglue_domain_rollouts": _export_superglue_domain_rollouts,
    }
    fn = handlers.get(entity)
    if not fn:
        raise HTTPException(status_code=400, detail=f"Unsupported entity: {entity}. Supported: {SUPPORTED_ENTITIES}")
    return fn(db, tenant_id, filt)


def _data_to_csv(data: list[dict]) -> io.BytesIO:
    if not data:
        buffer = io.BytesIO()
        w = csv.writer(io.TextIOWrapper(buffer, encoding="utf-8-sig", newline=""))
        w.writerow([])
        return buffer
    keys = list(data[0].keys())
    buffer = io.BytesIO()
    text_buffer = io.TextIOWrapper(buffer, encoding="utf-8-sig", newline="")
    writer = csv.DictWriter(text_buffer, fieldnames=keys)
    writer.writeheader()
    for row in data:
        writer.writerow(row)
    text_buffer.flush()
    buffer.seek(0)
    return buffer


def _data_to_xlsx(data: list[dict]) -> io.BytesIO:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise HTTPException(status_code=501, detail="Excel export requires openpyxl. Install with: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Export"
    if not data:
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    keys = list(data[0].keys())
    for col, key in enumerate(keys, 1):
        ws.cell(row=1, column=col, value=key)
    for row_idx, row in enumerate(data, 2):
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _filename(entity: str, fmt: str) -> str:
    today = date.today().isoformat()
    ext = "csv" if fmt == "csv" else "xlsx"
    return f"export_{entity}_{today}.{ext}"


@router.post("/list", response_class=StreamingResponse, summary="List exportieren",
    response_model=CompatFlexOut
)
async def export_list(
    body: ExportListRequest,
    tenant_id: str = Depends(get_tenant_id),
    db: Session = Depends(get_db),
):
    """
    Export entity list as CSV or XLSX (streaming).
    Request body: entity (required), format (csv|xlsx, default csv), filter (optional).
    """
    entity = (body.entity or "").strip().lower()
    if entity not in SUPPORTED_ENTITIES:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported entity: {entity}. Supported: {SUPPORTED_ENTITIES}",
        )
    fmt = (body.format or "csv").strip().lower()
    if fmt not in ("csv", "xlsx"):
        fmt = "csv"

    # Resolve tenant: filter can override for admin use; otherwise use request tenant
    effective_tenant = tenant_id
    if body.filter and body.filter.tenant_id:
        effective_tenant = body.filter.tenant_id

    data = _get_entity_data(entity, db, effective_tenant, body.filter)

    if fmt == "xlsx":
        buffer = _data_to_xlsx(data)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        buffer = _data_to_csv(data)
        media_type = "text/csv"

    filename = _filename(entity, fmt)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/entities", summary="Export entities auflisten",
    response_model=CompatFlexOut
)
async def list_export_entities():
    """Return supported entity values for export/list."""
    return {"entities": SUPPORTED_ENTITIES}
